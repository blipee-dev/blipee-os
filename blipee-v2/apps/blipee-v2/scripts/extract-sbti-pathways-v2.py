#!/usr/bin/env python3
"""
Extract SBTi pathway data from official Excel tools
Converts Database sheet to SQL INSERT statements
"""

import sys
import openpyxl

def extract_pathways_from_excel(excel_path, output_sql_path):
    """Extract pathway data from SBTi Excel Database sheet"""

    print(f"📖 Reading {excel_path}...")
    wb = openpyxl.load_workbook(excel_path, data_only=True)

    if 'Database' not in wb.sheetnames:
        print("❌ 'Database' sheet not found!")
        return

    ws = wb['Database']
    print(f"✅ Found Database sheet: {ws.max_row} rows × {ws.max_column} columns")

    # Map sector names to our enum
    sector_map = {
        'Iron and steel': 'iron_steel',
        'Cement': 'cement',
        'Aluminium': 'aluminum',
        'Aluminum': 'aluminum',
        'Pulp and paper': 'pulp_paper',
        'Other industry': 'cross_sector',
        'Services - Buildings': 'buildings',
        'Residential Buildings': 'buildings',
        'Power': 'power_generation',
        'Power generation': 'power_generation',
        'Transport': 'transport',
        'Primary energy demand and industry': 'cross_sector',
    }

    # Map scenario names
    scenario_map = {
        'ETP B2DS': 'ETP_B2DS',
        'SBTi 1.5C': 'SBTi_1.5C',
        'NZE2021': 'NZE2021',
    }

    # Read header to find year columns (col 9 onwards are years as strings '2014', '2015', etc)
    year_columns = []
    for col_idx in range(9, ws.max_column + 1):  # Years start at column 9
        header = ws.cell(1, col_idx).value
        if header and isinstance(header, str):
            try:
                year = int(header)
                if 2014 <= year <= 2050:
                    year_columns.append((col_idx, year))
            except ValueError:
                pass

    print(f"📅 Found {len(year_columns)} year columns: {[y for _, y in year_columns[:5]]} ... {[y for _, y in year_columns[-3:]]}")

    # Generate SQL INSERT statements
    sql_statements = []
    processed_rows = 0
    skipped_rows = 0

    for row_idx in range(2, ws.max_row + 1):  # Skip header
        # Extract metadata
        scenario_raw = ws.cell(row_idx, 2).value  # Col B: Scenario
        region = ws.cell(row_idx, 4).value or 'World'  # Col D: Region
        metric_type_raw = ws.cell(row_idx, 5).value  # Col E: Flow (Emissions/Electricity/etc)
        unit = ws.cell(row_idx, 6).value  # Col F: Unit
        sector_raw = ws.cell(row_idx, 8).value  # Col H: Sector

        # Map to our format
        scenario = scenario_map.get(scenario_raw)
        sector = sector_map.get(sector_raw)

        # Determine metric type
        if metric_type_raw == 'Emissions':
            metric_type = 'Emissions'
        elif metric_type_raw in ['Electricity', 'Activity']:
            metric_type = 'Activity'
        else:
            metric_type = metric_type_raw

        # Skip if we don't recognize sector or scenario
        if not sector or not scenario:
            skipped_rows += 1
            continue

        # Only process emissions data for now (activity data can be added later if needed)
        if metric_type != 'Emissions':
            continue

        # Extract year values
        values_found = 0
        for col_idx, year in year_columns:
            value = ws.cell(row_idx, col_idx).value

            if value is not None and isinstance(value, (int, float)) and value != 0:
                # Convert MtCO2 to tCO2 if needed (for consistency)
                # We'll keep original units and let the calculator handle scaling
                sql = (
                    f"INSERT INTO sbti_pathways (scenario, sector, region, metric_type, unit, year, value, data_source) "
                    f"VALUES ('{scenario}', '{sector}', '{region}', '{metric_type}', '{unit}', {year}, {value}, 'IEA SBTi Tool v2.4') "
                    f"ON CONFLICT (scenario, sector, region, metric_type, unit, year) DO NOTHING;"
                )
                sql_statements.append(sql)
                values_found += 1

        if values_found > 0:
            processed_rows += 1

    print(f"\n✅ Generated {len(sql_statements)} SQL INSERT statements from {processed_rows} data rows")
    print(f"⏭️  Skipped {skipped_rows} rows (unrecognized sector/scenario)")

    # Write to file
    with open(output_sql_path, 'w', encoding='utf-8') as f:
        f.write("-- ============================================================================\n")
        f.write("-- SBTI PATHWAYS DATA - Complete dataset from Official Excel Tools\n")
        f.write("-- ============================================================================\n")
        f.write("-- Source: SBTi Target-Setting Tool v2.4\n")
        f.write("-- Database sheet: 259 rows of IEA scenario data\n")
        f.write("-- Scenarios: ETP B2DS (Beyond 2°C), SBTi 1.5C, NZE2021\n")
        f.write("-- Sectors: Cross-sector, Iron & Steel, Cement, Aluminum, Power, etc.\n")
        f.write("-- Years: 2014-2050\n")
        f.write("-- ============================================================================\n\n")

        # Write inserts in batches for readability
        f.write("-- Delete existing sample data\n")
        f.write("DELETE FROM sbti_pathways WHERE data_source LIKE '%IEA%';\n\n")

        for i, sql in enumerate(sql_statements):
            f.write(sql + "\n")
            if (i + 1) % 50 == 0:
                f.write(f"\n-- Progress: {i + 1}/{len(sql_statements)} rows\n\n")

        f.write("\n-- ============================================================================\n")
        f.write("-- VERIFICATION QUERIES\n")
        f.write("-- ============================================================================\n\n")
        f.write("-- Count by scenario and sector\n")
        f.write("SELECT scenario, sector, COUNT(*) as row_count \n")
        f.write("FROM sbti_pathways \n")
        f.write("GROUP BY scenario, sector \n")
        f.write("ORDER BY scenario, sector;\n\n")

        f.write("-- Sample data for verification\n")
        f.write("SELECT * FROM sbti_pathways \n")
        f.write("WHERE sector = 'cement' AND scenario = 'SBTi_1.5C' \n")
        f.write("ORDER BY year \n")
        f.write("LIMIT 10;\n")

    print(f"💾 Saved to {output_sql_path}")
    print(f"\n📊 Summary:")
    print(f"   - Total SQL statements: {len(sql_statements)}")
    print(f"   - Processed rows: {processed_rows}")
    print(f"   - Years covered: {year_columns[0][1]} to {year_columns[-1][1]}")
    print(f"\n🚀 Next steps:")
    print(f"   1. Review: {output_sql_path}")
    print(f"   2. Apply migration to database")
    print(f"   3. Verify data imported correctly")


if __name__ == '__main__':
    if len(sys.argv) < 2:
        print("Usage: python extract-sbti-pathways-v2.py <path-to-excel-file>")
        print("\nExample:")
        print("  python extract-sbti-pathways-v2.py ~/Downloads/SBTi-target-setting-tool.xlsx")
        sys.exit(1)

    excel_path = sys.argv[1]
    output_sql_path = '/Users/pedro/Documents/blipee/blipee-os/blipee-os/blipee-v2/apps/blipee-v2/supabase/migrations/20250111_sbti_pathways_complete.sql'

    extract_pathways_from_excel(excel_path, output_sql_path)
