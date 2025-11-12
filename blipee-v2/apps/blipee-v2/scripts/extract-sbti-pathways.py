#!/usr/bin/env python3
"""
Extract SBTi pathway data from official Excel tools
Converts Database sheet to SQL INSERT statements
"""

import sys
import openpyxl
from pathlib import Path

def extract_pathways_from_excel(excel_path, output_sql_path):
    """Extract pathway data from SBTi Excel Database sheet"""

    print(f"📖 Reading {excel_path}...")
    wb = openpyxl.load_workbook(excel_path, data_only=True)

    # Database sheet contains the pathway data
    if 'Database' not in wb.sheetnames:
        print("❌ 'Database' sheet not found!")
        return

    ws = wb['Database']
    print(f"✅ Found Database sheet: {ws.max_row} rows × {ws.max_column} columns")

    # Read header row
    headers = []
    for cell in ws[1]:
        headers.append(cell.value)

    print(f"📊 Headers: {headers[:10]}...")  # Show first 10

    # Find year columns (typically start from column 9 onwards)
    year_columns = []
    for i, header in enumerate(headers):
        if isinstance(header, (int, float)) and 2014 <= header <= 2050:
            year_columns.append((i, int(header)))

    print(f"📅 Found {len(year_columns)} year columns: {[y for _, y in year_columns[:5]]}...")

    # Map sector names to our enum
    sector_map = {
        'Iron and steel': 'iron_steel',
        'Cement': 'cement',
        'Aluminium': 'aluminum',
        'Pulp and paper': 'pulp_paper',
        'Other industry': 'cross_sector',
        'Services - Buildings': 'buildings',
        'Power': 'power_generation',
        'Power generation': 'power_generation',
    }

    # Map scenario names
    scenario_map = {
        'ETP B2DS': 'ETP_B2DS',
        'SBTi 1.5C': 'SBTi_1.5C',
        'NZE2021': 'NZE2021',
    }

    # Generate SQL INSERT statements
    sql_statements = []
    row_count = 0

    for row_idx in range(2, ws.max_row + 1):  # Skip header
        row = list(ws[row_idx])

        # Extract metadata from first columns
        # Typical structure: Model.Parameter, Scenario, Sheet, Region, Flow.1, Unit, Table.1, Sector.ETP, then years...
        scenario_raw = row[1].value if len(row) > 1 else None
        region = row[3].value if len(row) > 3 else 'World'
        metric_type_raw = row[4].value if len(row) > 4 else None
        unit = row[5].value if len(row) > 5 else None
        sector_raw = row[7].value if len(row) > 7 else None

        # Map to our format
        scenario = scenario_map.get(scenario_raw, scenario_raw)
        sector = sector_map.get(sector_raw, None)

        # Determine metric type
        metric_type = 'Emissions' if metric_type_raw == 'Emissions' else 'Activity'

        # Skip if we don't recognize sector or scenario
        if not sector or not scenario or scenario not in scenario_map.values():
            continue

        # Extract year values
        for col_idx, year in year_columns:
            if col_idx < len(row):
                value = row[col_idx].value

                if value is not None and isinstance(value, (int, float)) and value != 0:
                    # Generate SQL INSERT
                    sql = f"INSERT INTO sbti_pathways (scenario, sector, region, metric_type, unit, year, value, data_source) VALUES ('{scenario}', '{sector}', '{region}', '{metric_type}', '{unit}', {year}, {value}, 'IEA Database Sheet');"
                    sql_statements.append(sql)
                    row_count += 1

    print(f"\n✅ Generated {row_count} SQL INSERT statements")

    # Write to file
    with open(output_sql_path, 'w') as f:
        f.write("-- ============================================================================\n")
        f.write("-- SBTI PATHWAYS DATA - Extracted from Official Excel Tools\n")
        f.write(f"-- Source: {Path(excel_path).name}\n")
        f.write("-- ============================================================================\n\n")
        f.write("-- Disable triggers for bulk insert\n")
        f.write("ALTER TABLE sbti_pathways DISABLE TRIGGER ALL;\n\n")

        for sql in sql_statements:
            f.write(sql + "\n")

        f.write("\n-- Re-enable triggers\n")
        f.write("ALTER TABLE sbti_pathways ENABLE TRIGGER ALL;\n\n")
        f.write("-- Verify data\n")
        f.write("SELECT scenario, sector, COUNT(*) as row_count FROM sbti_pathways GROUP BY scenario, sector ORDER BY scenario, sector;\n")

    print(f"💾 Saved to {output_sql_path}")
    print(f"\n📊 Summary:")
    print(f"   - Total rows: {row_count}")
    print(f"\n🚀 Next steps:")
    print(f"   1. Review the generated SQL file")
    print(f"   2. Run: psql < {output_sql_path}")
    print(f"   3. Verify data imported correctly")


if __name__ == '__main__':
    if len(sys.argv) < 2:
        print("Usage: python extract-sbti-pathways.py <path-to-excel-file>")
        print("\nExample:")
        print("  python extract-sbti-pathways.py ~/Downloads/SBTi-target-setting-tool.xlsx")
        sys.exit(1)

    excel_path = sys.argv[1]
    output_sql_path = '/Users/pedro/Documents/blipee/blipee-os/blipee-os/blipee-v2/apps/blipee-v2/supabase/migrations/20250111_sbti_pathways_complete.sql'

    extract_pathways_from_excel(excel_path, output_sql_path)
