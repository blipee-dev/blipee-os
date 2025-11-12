#!/usr/bin/env python3
import openpyxl
import sys

excel_path = sys.argv[1] if len(sys.argv) > 1 else '/Users/pedro/Downloads/SBTi-target-setting-tool.xlsx'

wb = openpyxl.load_workbook(excel_path, data_only=True)
ws = wb['Database']

print(f"Database sheet: {ws.max_row} rows × {ws.max_column} columns\n")

# Show first row (headers)
print("Headers (first 20 columns):")
for i in range(1, min(21, ws.max_column + 1)):
    cell = ws.cell(1, i)
    print(f"  Col {i:2d}: {repr(cell.value)} (type: {type(cell.value).__name__})")

print("\n" + "="*60)
print("Sample data rows (first 3):")
for row_num in range(2, 5):
    print(f"\nRow {row_num}:")
    for i in range(1, min(13, ws.max_column + 1)):
        cell = ws.cell(row_num, i)
        val = cell.value
        if val is not None:
            print(f"  Col {i:2d}: {repr(val)[:50]}")
