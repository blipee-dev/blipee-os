#!/usr/bin/env python3
import openpyxl

wb = openpyxl.load_workbook('/Users/pedro/Downloads/SBTi-target-setting-tool.xlsx', data_only=True)
ws = wb['Database']

# Get unique sectors and scenarios
sectors = set()
scenarios = set()

for row_idx in range(2, ws.max_row + 1):
    sector = ws.cell(row_idx, 8).value  # Col H
    scenario = ws.cell(row_idx, 2).value  # Col B
    if sector:
        sectors.add(sector)
    if scenario:
        scenarios.add(scenario)

print("📊 Unique SECTORS found in Database sheet:")
for s in sorted(sectors):
    print(f"   - {s}")

print(f"\n📊 Unique SCENARIOS found:")
for s in sorted(scenarios):
    print(f"   - {s}")

print(f"\n Total unique sectors: {len(sectors)}")
print(f" Total unique scenarios: {len(scenarios)}")
